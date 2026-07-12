# apps/data_engine/connectors/readers.py
"""Concrete enterprise connector implementations for the MAC Connector Framework.

Includes:
- `CSVConnector`: High-performance delimited text reader with BOM and streaming support.
- `JSONConnector`: Structured object and nested list reader.
- `ExcelConnector`: Spreadsheet reader architecture prepared for `openpyxl`.
- `SQLConnector`: Relational database reader (DB-API 2.0 compliant: psycopg, pyodbc, mysqlclient, cx_Oracle).
- `RESTConnector`: HTTP/REST API client with authentication injection, timeouts, and retries.
"""

import csv
import io
import json
import os
import time
import urllib.error
import urllib.request
from typing import Any, Dict, Iterator, List, Optional

from .base import BaseConnector
from .contracts import ConnectorConfig
from .datasource import DataSource
from .exceptions import (
    AuthenticationException,
    ConnectionFailedException,
    ConnectorException,
    InvalidConfigurationException,
    TimeoutException,
)
from .registry import ConnectorRegistry


# ==============================================================================
# 1. CSV Connector
# ==============================================================================

class CSVConnector(BaseConnector):
    """High-performance CSV and delimited file connector with streaming support."""

    def connect(self) -> None:
        source = self.config.get_param("source")
        if not source:
            raise InvalidConfigurationException("CSVConnector requires a 'source' parameter (file path or string buffer).")
        self._is_connected = True

    def disconnect(self) -> None:
        self._is_connected = False

    def test_connection(self) -> bool:
        try:
            source = self.config.get_param("source")
            if not source:
                return False
            if isinstance(source, str) and os.path.exists(source):
                return os.path.isfile(source) and os.access(source, os.R_OK)
            # If string buffer containing CSV content or file-like object
            return True
        except Exception:
            return False

    def _get_io_stream(self, source: Any) -> Any:
        if isinstance(source, str):
            if os.path.exists(source) and os.path.isfile(source):
                encoding = self.config.get_param("encoding", "utf-8-sig")
                return open(source, "r", encoding=encoding, newline="")
            # Treat string as raw CSV text buffer
            return io.StringIO(source)
        if hasattr(source, "read"):
            return source
        raise ConnectionFailedException("Invalid source for CSVConnector: not a valid path, buffer, or string.")

    def fetch(
        self,
        query_or_path: Optional[str] = None,
        limit: Optional[int] = None,
        **kwargs: Any,
    ) -> List[Dict[str, Any]]:
        if not self.is_connected:
            self.connect()

        source = query_or_path or self.config.get_param("source")
        delimiter = self.config.get_param("delimiter", ",")
        quotechar = self.config.get_param("quotechar", '"')

        stream = self._get_io_stream(source)
        should_close = isinstance(source, str) and os.path.exists(source) and os.path.isfile(source)

        try:
            reader = csv.DictReader(stream, delimiter=delimiter, quotechar=quotechar)
            records: List[Dict[str, Any]] = []
            for row in reader:
                records.append(dict(row))
                if limit and len(records) >= limit:
                    break
            return records
        except Exception as exc:
            raise ConnectorException(f"Failed reading CSV data: {exc}") from exc
        finally:
            if should_close:
                stream.close()

    def stream(
        self,
        query_or_path: Optional[str] = None,
        chunk_size: int = 1000,
        **kwargs: Any,
    ) -> Iterator[List[Dict[str, Any]]]:
        if not self.is_connected:
            self.connect()

        source = query_or_path or self.config.get_param("source")
        delimiter = self.config.get_param("delimiter", ",")
        quotechar = self.config.get_param("quotechar", '"')

        stream = self._get_io_stream(source)
        should_close = isinstance(source, str) and os.path.exists(source) and os.path.isfile(source)

        try:
            reader = csv.DictReader(stream, delimiter=delimiter, quotechar=quotechar)
            batch: List[Dict[str, Any]] = []
            for row in reader:
                batch.append(dict(row))
                if len(batch) >= chunk_size:
                    yield batch
                    batch = []
            if batch:
                yield batch
        finally:
            if should_close:
                stream.close()

    def metadata(self) -> DataSource:
        source = self.config.get_param("source")
        columns: List[str] = []
        size_bytes = None

        if isinstance(source, str) and os.path.exists(source) and os.path.isfile(source):
            size_bytes = os.path.getsize(source)
            with open(source, "r", encoding=self.config.get_param("encoding", "utf-8-sig")) as f:
                header_line = f.readline()
                if header_line:
                    columns = [c.strip('" \r\n') for c in header_line.split(self.config.get_param("delimiter", ","))]
        elif isinstance(source, str):
            size_bytes = len(source.encode("utf-8"))
            lines = source.splitlines()
            if lines:
                columns = [c.strip('" ') for c in lines[0].split(self.config.get_param("delimiter", ","))]

        return DataSource(
            name=os.path.basename(source) if isinstance(source, str) and os.path.exists(source) else "csv_stream",
            source_type="csv",
            columns=columns,
            encoding=self.config.get_param("encoding", "utf-8"),
            dialect="delimited",
            size_bytes=size_bytes,
        )


# ==============================================================================
# 2. JSON Connector
# ==============================================================================

class JSONConnector(BaseConnector):
    """Structured JSON data source connector supporting nested extraction paths."""

    def connect(self) -> None:
        source = self.config.get_param("source")
        if not source:
            raise InvalidConfigurationException("JSONConnector requires a 'source' parameter (file path or JSON string/dict).")
        self._is_connected = True

    def disconnect(self) -> None:
        self._is_connected = False

    def test_connection(self) -> bool:
        try:
            self._load_raw_json(self.config.get_param("source"))
            return True
        except Exception:
            return False

    def _load_raw_json(self, source: Any) -> Any:
        if isinstance(source, (dict, list)):
            return source
        if isinstance(source, str):
            if os.path.exists(source) and os.path.isfile(source):
                with open(source, "r", encoding="utf-8") as f:
                    return json.load(f)
            return json.loads(source)
        raise ConnectionFailedException("Unsupported source type for JSONConnector.")

    def _extract_rows(self, raw_data: Any, root_path: Optional[str] = None) -> List[Dict[str, Any]]:
        target = raw_data
        if root_path:
            for key in root_path.split("."):
                if isinstance(target, dict) and key in target:
                    target = target[key]
                else:
                    target = []
                    break

        if isinstance(target, list):
            return [item if isinstance(item, dict) else {"value": item} for item in target]
        if isinstance(target, dict):
            return [target]
        return [{"value": target}]

    def fetch(
        self,
        query_or_path: Optional[str] = None,
        limit: Optional[int] = None,
        **kwargs: Any,
    ) -> List[Dict[str, Any]]:
        if not self.is_connected:
            self.connect()

        source = query_or_path or self.config.get_param("source")
        raw = self._load_raw_json(source)
        rows = self._extract_rows(raw, root_path=self.config.get_param("root_path"))
        return rows[:limit] if limit else rows

    def stream(
        self,
        query_or_path: Optional[str] = None,
        chunk_size: int = 1000,
        **kwargs: Any,
    ) -> Iterator[List[Dict[str, Any]]]:
        rows = self.fetch(query_or_path=query_or_path, **kwargs)
        for i in range(0, len(rows), chunk_size):
            yield rows[i : i + chunk_size]

    def metadata(self) -> DataSource:
        rows = self.fetch(limit=1)
        columns = list(rows[0].keys()) if rows else []
        return DataSource(
            name="json_source",
            source_type="json",
            columns=columns,
            encoding="utf-8",
            dialect=self.config.get_param("root_path", "root"),
        )


# ==============================================================================
# 3. Excel Connector
# ==============================================================================

class ExcelConnector(BaseConnector):
    """Excel spreadsheet connector prepared for openpyxl integration."""

    def connect(self) -> None:
        source = self.config.get_param("source")
        if not source:
            raise InvalidConfigurationException("ExcelConnector requires a 'source' parameter (.xlsx file or data rows).")
        self._is_connected = True

    def disconnect(self) -> None:
        self._is_connected = False

    def test_connection(self) -> bool:
        try:
            source = self.config.get_param("source")
            if isinstance(source, list):
                return True
            if isinstance(source, str) and os.path.exists(source) and source.endswith(".xlsx"):
                return os.access(source, os.R_OK)
            return True if source else False
        except Exception:
            return False

    def fetch(
        self,
        query_or_path: Optional[str] = None,
        limit: Optional[int] = None,
        **kwargs: Any,
    ) -> List[Dict[str, Any]]:
        if not self.is_connected:
            self.connect()

        source = query_or_path or self.config.get_param("source")
        # Direct rows simulation (Zero-ORM test mode / preloaded sheet rows)
        if isinstance(source, list):
            return source[:limit] if limit else source

        if isinstance(source, str) and os.path.exists(source):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(source, read_only=True)
                sheet_name = self.config.get_param("sheet_name") or wb.sheetnames[0]
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                wb.close()
                if not rows:
                    return []
                headers = [str(c) if c is not None else f"col_{idx}" for idx, c in enumerate(rows[0])]
                records = [dict(zip(headers, row)) for row in rows[1:]]
                return records[:limit] if limit else records
            except ImportError:
                raise UnsupportedConnectorException("openpyxl library is required to read real .xlsx files.")

        raise ConnectionFailedException("Unable to read Excel data: invalid source.")

    def stream(
        self,
        query_or_path: Optional[str] = None,
        chunk_size: int = 1000,
        **kwargs: Any,
    ) -> Iterator[List[Dict[str, Any]]]:
        rows = self.fetch(query_or_path=query_or_path, **kwargs)
        for i in range(0, len(rows), chunk_size):
            yield rows[i : i + chunk_size]

    def metadata(self) -> DataSource:
        rows = self.fetch(limit=1)
        columns = list(rows[0].keys()) if rows else []
        return DataSource(
            name=str(self.config.get_param("sheet_name", "Sheet1")),
            source_type="excel",
            columns=columns,
            dialect="openpyxl",
        )


# ==============================================================================
# 4. SQL Connector (DB-API 2.0 / Drivers)
# ==============================================================================

class SQLConnector(BaseConnector):
    """Relational database connector adhering to DB-API 2.0 without Django ORM."""

    def connect(self) -> None:
        if not self.config.host and not self.config.get_param("connection_object") and not self.config.get_param("mock_rows"):
            raise InvalidConfigurationException("SQLConnector requires host/connection parameters or mock object.")
        self._is_connected = True

    def disconnect(self) -> None:
        conn = self.config.get_param("connection_object")
        if conn and hasattr(conn, "close"):
            try:
                conn.close()
            except Exception:
                pass
        self._is_connected = False

    def test_connection(self) -> bool:
        try:
            if not self.is_connected:
                self.connect()
            conn = self.config.get_param("connection_object")
            if conn and hasattr(conn, "cursor"):
                cur = conn.cursor()
                cur.execute("SELECT 1")
                return True
            if self.config.get_param("mock_rows") is not None:
                return True
            return bool(self.config.host)
        except Exception:
            return False

    def fetch(
        self,
        query_or_path: Optional[str] = None,
        limit: Optional[int] = None,
        **kwargs: Any,
    ) -> List[Dict[str, Any]]:
        if not self.is_connected:
            self.connect()

        # Zero-ORM deterministic testing / mock execution
        mock_rows = self.config.get_param("mock_rows")
        if mock_rows is not None:
            return mock_rows[:limit] if limit else mock_rows

        conn = self.config.get_param("connection_object")
        query = query_or_path or self.config.get_param("query") or "SELECT * FROM data"
        if limit:
            query = f"{query} LIMIT {limit}"

        if conn and hasattr(conn, "cursor"):
            cur = conn.cursor()
            cur.execute(query)
            col_names = [desc[0] for desc in cur.description] if cur.description else []
            rows = cur.fetchall()
            return [dict(zip(col_names, row)) for row in rows]

        raise ConnectionFailedException("No active SQL DB-API 2.0 connection object available.")

    def stream(
        self,
        query_or_path: Optional[str] = None,
        chunk_size: int = 1000,
        **kwargs: Any,
    ) -> Iterator[List[Dict[str, Any]]]:
        mock_rows = self.config.get_param("mock_rows")
        if mock_rows is not None:
            for i in range(0, len(mock_rows), chunk_size):
                yield mock_rows[i : i + chunk_size]
            return

        conn = self.config.get_param("connection_object")
        query = query_or_path or self.config.get_param("query") or "SELECT * FROM data"
        if conn and hasattr(conn, "cursor"):
            cur = conn.cursor()
            cur.execute(query)
            col_names = [desc[0] for desc in cur.description] if cur.description else []
            while True:
                batch = cur.fetchmany(chunk_size)
                if not batch:
                    break
                yield [dict(zip(col_names, row)) for row in batch]
            return

        raise ConnectionFailedException("No active SQL DB-API 2.0 connection object for streaming.")

    def metadata(self) -> DataSource:
        rows = self.fetch(limit=1)
        columns = list(rows[0].keys()) if rows else []
        return DataSource(
            name=str(self.config.get_param("table", "sql_query")),
            source_type="sql",
            columns=columns,
            dialect=self.config.get_param("driver", "generic_sql"),
        )


# ==============================================================================
# 5. REST Connector
# ==============================================================================

class RESTConnector(BaseConnector):
    """HTTP/REST API connector with authentication injection, timeout, and retry policies."""

    def connect(self) -> None:
        url = self.config.get_param("url") or self.config.host
        if not url:
            raise InvalidConfigurationException("RESTConnector requires 'url' or 'host' parameter.")
        self._is_connected = True

    def disconnect(self) -> None:
        self._is_connected = False

    def test_connection(self) -> bool:
        try:
            self.fetch(limit=1)
            return True
        except Exception:
            return False

    def fetch(
        self,
        query_or_path: Optional[str] = None,
        limit: Optional[int] = None,
        **kwargs: Any,
    ) -> List[Dict[str, Any]]:
        if not self.is_connected:
            self.connect()

        url = query_or_path or self.config.get_param("url") or self.config.host
        method = self.config.get_param("method", "GET").upper()
        headers = dict(self.config.get_param("headers", {}))
        max_retries = int(self.config.get_param("max_retries", 3))
        backoff = float(self.config.get_param("retry_backoff", 0.1))

        if self.config.auth_provider:
            headers = self.config.auth_provider.apply_auth(headers)

        # Allow transport callback injection for isolated testing without live sockets
        transport_cb = self.config.get_param("transport_callback")
        if transport_cb:
            raw_res = transport_cb(method=method, url=url, headers=headers)
            if isinstance(raw_res, list):
                return raw_res[:limit] if limit else raw_res
            if isinstance(raw_res, dict):
                return [raw_res]

        for attempt in range(max_retries + 1):
            try:
                req = urllib.request.Request(url, method=method, headers=headers)
                with urllib.request.urlopen(req, timeout=self.config.timeout) as resp:
                    status_code = resp.getcode()
                    if status_code in (401, 403):
                        raise AuthenticationException(f"API authentication rejected ({status_code}).")
                    body = resp.read().decode("utf-8")
                    parsed = json.loads(body)
                    if isinstance(parsed, list):
                        records = [item if isinstance(item, dict) else {"value": item} for item in parsed]
                    elif isinstance(parsed, dict) and "results" in parsed and isinstance(parsed["results"], list):
                        records = parsed["results"]
                    elif isinstance(parsed, dict):
                        records = [parsed]
                    else:
                        records = [{"value": parsed}]
                    return records[:limit] if limit else records
            except urllib.error.HTTPError as err:
                if err.code in (401, 403):
                    raise AuthenticationException(f"API authentication rejected: {err}") from err
                if err.code in (500, 502, 503, 504) and attempt < max_retries:
                    time.sleep(backoff * (2 ** attempt))
                    continue
                raise ConnectorException(f"HTTP error during REST fetch: {err}") from err
            except (urllib.error.URLError, TimeoutError) as err:
                if attempt < max_retries:
                    time.sleep(backoff * (2 ** attempt))
                    continue
                raise TimeoutException(f"Network timeout reaching REST endpoint {url}: {err}") from err

        raise ConnectorException(f"Failed reaching REST endpoint {url} after {max_retries} retries.")

    def stream(
        self,
        query_or_path: Optional[str] = None,
        chunk_size: int = 1000,
        **kwargs: Any,
    ) -> Iterator[List[Dict[str, Any]]]:
        rows = self.fetch(query_or_path=query_or_path, **kwargs)
        for i in range(0, len(rows), chunk_size):
            yield rows[i : i + chunk_size]

    def metadata(self) -> DataSource:
        url = self.config.get_param("url") or self.config.host
        rows = self.fetch(limit=1)
        columns = list(rows[0].keys()) if rows else []
        return DataSource(
            name=str(url),
            source_type="rest",
            columns=columns,
            dialect=self.config.get_param("method", "GET"),
        )


# Register built-in enterprise connectors into global singleton
ConnectorRegistry.global_registry().register("csv", CSVConnector)
ConnectorRegistry.global_registry().register("json", JSONConnector)
ConnectorRegistry.global_registry().register("excel", ExcelConnector)
ConnectorRegistry.global_registry().register("sql", SQLConnector)
ConnectorRegistry.global_registry().register("rest", RESTConnector)
